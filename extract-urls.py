#! /usr/bin/python3

from pyxlsb import open_workbook
from colorama import Fore, Style
from openpyxl import load_workbook #
import xlrd #
import pandas as pd #
import os
import sys
import optparse
import requests
import re
import json #
import subprocess #
import msoffcrypto
import io
from tempfile import NamedTemporaryFile

URLHAUS_API = "https://urlhaus-api.abuse.ch/v1/url/"
regex_uri = "https?:\/\/[a-zA-Z0-9\.\/\-\:]{5,}\.\w{1,}"

def setup_args():
	parser = optparse.OptionParser()

	parser.add_option('-d', '--directory',
	action="store", dest="directory",
	help="The folder that contains your Dridex docs", default=".")

	return parser.parse_args()

def extract_downloader_links(file_path):
	try:
		try: # first let us try to open the file in xlrd 
			with open(file_path, "rb") as f, NamedTemporaryFile(suffix='.xls', delete=False) as tmpf:
					file = msoffcrypto.OfficeFile(f)
					if file_path[-12:] == 'dridex01.xls':
						file.load_key(password='50821')
					elif file_path[-12:] == 'dridex02.xls':
						file.load_key(password='68443')
					elif file_path[-12:] == 'dridex03.xls':
						file.load_key(password='27158')
					else:
						file.load_key(password='VelvetSweatshop')
					file.decrypt(tmpf)

			xls = xlrd.open_workbook(tmpf.name)
			print(f"{Fore.GREEN}[*]{Style.RESET_ALL} Working file: " + file_path)
			obf = []

			# for every sheet we run through every row and coll and look at the cell value if its not empty
			for sheet in xls.sheets():
				ws = xls[sheet.name]
			
				rows = ws.nrows
				cols = ws.ncols
				# if the cell is not empty we append it to obf just like the original script the 
				# rest of the script can then pick through the info to extract the urls
				for row in range(rows):
					#obf.append(sheet.row_values(rowx=row, start_colx=0, end_colx=None))
					for col in range(cols):
						cell = sheet.cell(row,col)
						if cell.value!=None and cell.value!="":
							obf.append(cell.value)
							# print(cell.value, end = ' ')  # uncomment this code to see the values from the encrypted workbooks
				
		except:	
			#try: # if we couldn't open it with xlrd then we will try openpyxl
								
				xls = load_workbook(filename=file_path, data_only=True)
				print(f"{Fore.GREEN}[*]{Style.RESET_ALL} Working file: " + file_path)
				obf = []

				for sheet in xls.sheetnames:
					ws = xls[sheet]

					for row in ws.rows:
						for c in row:
							if c.value!=None:
								obf.append(c.value)
			
		urls = []
		script = ""
		tmp_urls = ""

		if len(obf) == 2:
			for y in range(len(obf[1])):
				tmp_urls = tmp_urls + chr(ord(obf[1][y]) + int(obf[0][y]))

			if tmp_urls:
				url_string = tmp_urls.split("RSab")
				for url in url_string[0].split("E,"):
					urls.append("https://" + url)
		else:
			for y in range(len(obf)):
				if isinstance(obf[y], float):
					script = script + chr(int(obf[y]))

			if script:
				tmp_urls = re.findall(regex_uri, script)
				for url in tmp_urls:
					urls.append(url)

		print(f"{Fore.LIGHTYELLOW_EX}[+]{Style.RESET_ALL} Found " + str(len(urls)) + " URLs")
		for url in urls:
				signature_found = False
				print("\t[$] Found - " + url)
				
				# we make a request to URLHaus and see what the response is
				url_haus = 'https://urlhaus.abuse.ch/api/' ###
				api_key = 'af5b41988241d8a66d40' 
				
				jsonData = {
					'token' : api_key, 'anonymous' : '0', 'submission' : [
						{
							'url' : url, 'threat': 'malware_download'
						}
					]	
				
				}
				headers = { 'Content-Type' : 'application/json'}
				r = requests.post(url_haus, json=jsonData, timeout=15, headers=headers)
				
				# if the response has the word already in it, then its known 
				# if it has the word que in it, it has to still be processed but its been submitted
				if 'already' in str(r.content):
					if 'que' in str(r.content):
						print("\t\tAlready queued")
					else:
						print("\t\tAlready known")
				else:
					print("\t\tURL is new")
				output = subprocess.Popen("wget -O- --post-data='url="+url+"' https://urlhaus-api.abuse.ch/v1/url/", shell=True, stderr=subprocess.DEVNULL, stdout=subprocess.PIPE)
				outcome = output.stdout.read().split()
				# here we look for the results of the signature from the response of our wget to URLHaus
				if not 'no_results' in str(outcome[2]):
					for i in range(len(outcome)):
						if 'signature' in str(outcome[i]):
							print("\t\tURL Signature is: "+str(outcome[i+1])[3:-3])
							signature_found = True
				# if we find no signature then say so
				if not signature_found:
					print("\t\tNo signature has been given to this url.") ###
	except:
			print(f"{Fore.RED}[-] Error:{Style.RESET_ALL} could not open " + file_path)

	

def main(argv):
	options, args = setup_args() 

	for file_name in os.listdir(options.directory):
		extract_downloader_links(options.directory + "/"+file_name)

if __name__ == '__main__':
	main(sys.argv[1:])